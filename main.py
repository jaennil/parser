from utils import categories
import time
import count_words
import school_names
import query2links
import init_browser
import link2text

from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook
import json


def main():
    start = time.time()
    browser = init_browser.default()
    try:
        school_names_list = school_names.get()
        # school_names_list = ["Центральная музыкальная школа при Московской консерватории", "Московская средняя специальная музыкальная школа имени Гнесиных",             "Московская центральная художественная школа при Российской академии художеств(МЦХШ РАХ)"]
        schools = {}
        with open("words_removed.json", "r") as f:
            my_dict = json.load(f)
        cnt = 1
        for name in school_names_list:
            print(cnt/len(school_names_list)*100)
            schools[name] = {}
            schools[name]["points"] = {}
            # schools[name]["points"][category] = 0
            links = query2links.get(name, browser, amount=1)
            schools[name]["links"] = links
            for link in links:
                text = link2text.get_page_text(link, browser)
                words = count_words.count_words(text)
                for word_dict in words:
                    word = [key for key in word_dict.keys()][0]
                    for category in categories:
                        for word_dict_json in my_dict[category]["words"]:
                            json_word = [key for key in word_dict_json.keys()][0]
                            if word == json_word:
                                if schools[name]["points"].get(category) is not None:
                                    schools[name]["points"][category] += word_dict[word]
                                else:
                                    schools[name]["points"][category] = 0
                                break
            cnt += 1
        start = 3
        wb = load_workbook("./dataset.xlsx")
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[0]]
        for school in schools:
            # sort dict
            schools[school]["points"] = dict(sorted(schools[school]["points"].items(), key=lambda item: item[1], reverse=True))
            max_cat = list(schools[school]["points"].keys())[0]
            Sheet1 .cell(row = start, column = 24).value = max_cat #This will change the cell(2,4) to 4
            start += 1
        wb.save("dataset_categories.xlsx")
    except Exception:
        # throw exception
        raise Exception


    finally:
        browser.quit()
    print(f"Time: {time.time() - start}")


if __name__ == '__main__':
    main()
